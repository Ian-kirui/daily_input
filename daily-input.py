from tkinter import * 
import tkinter as tk
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

main=tk.Tk()
main.title("input daily data")
main.iconbitmap("I:\data_diary\diary.ico")
main.geometry("500x700")
main.config(highlightbackground= "black" , highlightthickness=2)

def submit():
    
    q = date.get()
    r = dayrating.get()
    s = dayoftheweek.get()
    t = weather.get()
    u = people.get()
    v = hoursofsleep.get()
    w = food.get()
    x = workout.get()
    y = activities.get()
    z = amountspent.get()

    print(q)
    print(r)
    print(s)
    print(t)
    print(u)
    print(v)
    print(w)
    print(x)
    print(y)
    print(z)

    file = pathlib.Path("data-diary.xlsx")
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet["A1"]="Date"
        sheet["B1"]="Day Rating"
        sheet["C1"]="Day of the week"
        sheet["D1"]="Weather"
        sheet["E1"]="People"
        sheet["F1"]="Hours of sleep"
        sheet["G1"]="Food"
        sheet["H1"]="Workout"
        sheet["I1"]="Activities"
        sheet["J1"]="Amount spent(ksh)"

        file.save("data-diary.xlsx")

    #creating workbook/dataset
    file=openpyxl.load_workbook("data-diary.xlsx")
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1, value = q)
    sheet.cell(column=2, row=sheet.max_row, value = r)
    sheet.cell(column=3, row=sheet.max_row, value = s)
    sheet.cell(column=4, row=sheet.max_row, value=t )
    sheet.cell(column=5, row=sheet.max_row, value=u )
    sheet.cell(column=6, row=sheet.max_row, value=v )
    sheet.cell(column=7, row=sheet.max_row, value=w )
    sheet.cell(column=8, row=sheet.max_row, value=x )
    sheet.cell(column=9, row=sheet.max_row, value=y )
    sheet.cell(column=10, row=sheet.max_row, value=z )

    file.save("data-diary.xlsx")





#creating a backup for the user
#xlfile = pd.read_excel('weather_data.xlsx', 'Sheet') #reading the excel file
#xlfile.to_csv('weather_data.csv', index=False)#conversion to csv



#building the frame 
frame = LabelFrame(main, text ='input weather data:').pack()

Label(frame, text= "Date(MM/DD/YYYY):").place(x=50, y=30)
Label(frame, text= "Day rating(good or bad):").place(x=50, y=70)
Label(frame, text= "Day of the week:").place(x=50, y=110)
Label(frame, text= "Weather:").place(x=50, y=150)
Label(frame, text= "People:").place(x=50, y=190)
Label(frame, text= "Hours of sleep(number only):").place(x=50, y=230)
Label(frame, text= "Food:").place(x=50, y=270)
Label(frame, text= "Workout(yes or no):").place(x=50, y=310)
Label(frame, text= "Activities:").place(x=50, y=350)
Label(main, text= "Amount spent(ksh):").place(x=50 , y=390)

date= Entry(frame)
date.place(x=250, y=30)

dayrating = Entry(frame)
dayrating.place(x=250, y=70)

dayoftheweek = Entry(frame)
dayoftheweek.place(x=250, y= 110)

weather = Entry(frame)
weather.place(x=250 ,y= 150)

people = Entry(frame)
people.place(x=250, y=190)

hoursofsleep = Entry(frame)
hoursofsleep.place(x=250, y=230)

food = Entry(frame)
food.place(x=250, y=270)

workout = Entry(frame)
workout.place(x=250, y=310)

activities = Entry(frame)
activities.place(x=250, y=350)

amountspent = Entry(frame)
amountspent.place(x=250, y=390)

button = tk.Button(main, text="submit" ,command= submit )
button.place(x=250, y= 420)


main.mainloop()

